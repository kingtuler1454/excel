
from openpyxl import Workbook
import openpyxl as op
from tqdm import trange  # для отображения прогресса в консоли
from progress.bar import IncrementalBar  # для отображения прогресса в консоли
from os import listdir, getcwd,path
import shutil


def read_cicle(filename: str, Solder:bool): # 44 столбца заранее известных

    information = []  # список информации о студентах
    column_names = [
        "ФГОО ВО в котором обучается студент",#1
        "ФГОО ВО при которой создан ВУЦ",#2
        "ВУС",#3 
        "ОВУ",#4
        "ФГОС",#5
        "Программа военной подготовки",#6
        "Год зачисления в ВУЗ",#7
        "Год начала обучения в ВУЦ",#8
        "Месяц нала обучения в ВУЦ",#9
        "Срок обучения (месяцев)",#10
        "Год окончания ВУЦ",#11
        "Месяц окончания обучения в ВУЦ",#12
        "Отметка о завершении военной подготовки",#13
        "Год окончания ВУЗа",#14
        "Месяц окончания обучения в ВУЗе",#15
        "Личный номер",#16
        "Фамилия",#17
        "Имя",#18
        "Отчество",#19
        "Дата рождения",#20
        "Место рождения",#21
        "Национальность",#22
        "Пол",#23
        "СНИЛС",#24
        "ИНН",#25
        "серия1",#26
        "серия2",#27
        "Номер",#28
        "Выдан",#29
        "Код подразделения",#30
        "Кем выдан",#31
        "Адрес регистрации",#32
        "Номер телефона",#33
        "Семейное положение",#34
        "Количество детей",#35
        "Военный комиссариат (по месту жительства)",#36
        "Военный комиссариат (в который будет направлено личное дело)",#37
        "По призыву",#38
        "По контракту",#39
        "Примечание",#40
        "Статус",	#41
        "Приказ о зачислении",	#42
        "Приказ о отчислении",	#43
        "Причина отчисления"    #44																							
    ]
    if Solder: column_names.insert(2,"Код должности")
    wb = op.load_workbook(filename, data_only=True)
    sheet = wb.active
    bar = IncrementalBar( " ", max=sheet.max_row -4 ) 
    for student_number in range (5,sheet.max_row + 1):

        column_students = []  # список значений студента в этих столбцах
        for i in range(1,sheet.max_column + 1):column_students.append(sheet.cell(row=student_number, column=i).value )
        tmp=dict(zip(column_names, column_students))
        if Solder and str(tmp['Код должности'])!='256':
            tmp_vuc=tmp['Код должности']
            tmp['Код должности']='256' 
            tmp['ВУС']=tmp_vuc

        information.append(tmp)  # создаем из списка имён столбцов и список значений студента словарик  и добавляем каждый словарик в список information  
        
        bar.next()
    return information

def read_excel(filename: str):  # функция принимающая имя эксель таблицы
    wb = op.load_workbook(filename, data_only=True)
    sheet = wb.active
    if sheet.cell(row=1, column=4).value=="Код должности":
        return read_cicle(filename,True)
    if sheet.cell(row=1, column=4).value=="ОВУ":
        return read_cicle(filename,False)

    information = []  # список информации о студентах
    column_names = []  # список имён столбцов
    for i in range(1, sheet.max_column + 1): column_names.append(sheet.cell(row=1, column=i).value)
    
    
    for student_number in trange(2, sheet.max_row + 1):  # бежим по каждой строчке
        column_students = []  # список значений студента в этих столбцах
        for i in range(1, sheet.max_column + 1):  # бежим по каждому столбцу
            column_students.append(sheet.cell(row=student_number, column=i).value)  # добавляем в список значений студента текущее значение студента
        tmp=dict(zip(column_names, column_students))
        for elem in tmp.values():
                if str(elem)[:5]=='солда' or str(elem)[:5]=='рядов' :
                    if 'Код должности' in tmp.keys()  and 'ВУС' in tmp.keys() !='256':
                        tmp_vuc=tmp['Код должности']
                        tmp['Код должности']='256' 
                        tmp['ВУС']=tmp_vuc
                        
        information.append(tmp)  # создаем из списка имён столбцов и список значений студента словарик  и добавляем каждый словарик в список information
    return information  # вовзращаем список из словарей

def sum(list1, list2):  # функция сложения двух списков из двух эксель таблиц

    list1=uniq_list(list1)
    list2=uniq_list(list2)
    bar = IncrementalBar( " ", max=len( list1) ) 
    
    tmp2=list2
    for elem1 in list1:  # для каждого элемента из первой таблицы
        for elem2 in list2:  # для каждого элемента из второйтаблицы
            # теперь проверяем есть ли среди двух таблиц повторяющийся студент

                if str(elem1.get("Фамилия")).lower().strip() == str(elem2.get("Фамилия")).lower().strip():  # если даты равны то затем фамилии,
                    if str(elem1.get("Отчество")).lower().strip() == str(elem2.get("Отчество")).lower().strip():
                        if str(elem1.get("Имя")).lower().strip() == str(elem2.get("Имя")).lower().strip():  # ну если и имена равны то это явно один и тот же человек
                            value = set(elem2) - set(elem1)  # смотрим разницу между словарями. получитс список ключей второго словаря, которых нет в первом
                            for new_key in list(value):  # для каждого такого ключа добавляем занчение  в первый словарь
                                elem1[new_key] = elem2.get(new_key)
                            tmp2.remove(elem2) #удаляем значения одинаковых студентов из копии списка2
        bar.next()  # для отображения прогресса в консоли
    return list1+tmp2  # возвращаем дополненный первый список

def uniq_list(information):
    list1=[]
    for elem in information:
        tmp={key:val for key,val in elem.items() if val != None  and val!=''}
        if (len(tmp))!=0:list1.append(tmp) # удаляем значения у словарей где None 

    tmp=[] #для индексов дубликатов
    for i in range (len(list1)-1):
        for j in range (i+1,len(list1)-2):

                if str(list1[i].get("Имя")).lower().strip()==str(list1[j].get("Имя")).lower().strip():
                    if str(list1[i].get("Фамилия")).lower().strip()== str(list1[j].get("Фамилия")).lower().strip():
                        if str(list1[i].get("Отчество")).lower().strip()== str(list1[j].get("Отчество")).lower().strip():
                            if len(list1[i])<=len(list1[j]):
                                value = set(list1[j]) - set(list1[i])  # смотрим разницу между словарями. получитс список ключей второго словаря, которых нет в первом
                                for new_key in list(value):  # для каждого такого ключа добавляем занчение  в первый словарь
                                    list1[i][new_key] = list1[j].get(new_key)
                            tmp.append(j) 
    tmp=sorted(tmp)

    for i in range(len(tmp),0):
        list1.pop(tmp[i]-i)

    return list1
#для суммирования вспомогаюзая

def write_excel(information,number):
    if number==1:
        name_column=[ 
        "ФГОО ВО в котором обучается студент",#1
        "ФГОО ВО при которой создан ВУЦ",#2
        "ВУС",#3 
        "ОВУ",#4
        "ФГОС",#5
        "Программа военной подготовки",#6
        "Год зачисления в ВУЗ",#7
        "Год начала обучения в ВУЦ",#8
        "Месяц нала обучения в ВУЦ",#9
        "Срок обучения (месяцев)",#10
        "Год окончания ВУЦ",#11
        "Месяц окончания обучения в ВУЦ",#12
        "Отметка о завершении военной подготовки",#13
        "Год окончания ВУЗа",#14
        "Месяц окончания обучения в ВУЗе",#15
        "Личный номер",#16
        "Фамилия",#17
        "Имя",#18
        "Отчество",#19
        "Дата рождения",#20
        "Место рождения",#21
        "Национальность",#22
        "Пол",#23
        "СНИЛС",#24
        "ИНН",#25
        "серия1",#26
        "серия2",#27
        "Номер",#28
        "Выдан",#29
        "Код подразделения",#30
        "Кем выдан",#31
        "Адрес регистрации",#32
        "Номер телефона",#33
        "Семейное положение",#34
        "Количество детей",#35
        "Военный комиссариат (по месту жительства)",#36
        "Военный комиссариат (в который будет направлено личное дело)",#37
        "По призыву",#38
        "По контракту",#39
        "Примечание",#40
        "Статус",	#41
        "Приказ о зачислении",	#42
        "Приказ о отчислении",	#43
        "Причина отчисления"    #44			
        ]
        shutil.copyfile('workprogram/'+'Не трогать!(Шаблон офицеров).xlsx', 'Итоговые таблицы/(Итог)Офицеры.xlsx')
        wb = op.load_workbook('Итоговые таблицы/(Итог)Офицеры.xlsx', data_only=True)
        sheet = wb.active
        id=0
        for j in range (len( information)):
            vuc=str(information[j].get('ВУС'))            
            program_pdgotovki=str(information[j].get('Программа военной подготовки'))
            if  program_pdgotovki[:2]=='оф' or vuc=='461000' or vuc=='461100' or vuc=='461200' or vuc=='461300'  :
                        
                            for i in range (1,len(name_column)+1):
                                tmp=''
                                if  name_column[i-1] in information[j].keys(): tmp=information[j][name_column[i-1]]
                                sheet.cell(row=id+5, column=i, value=tmp)
                            id+=1
        wb.save('Итоговые таблицы/(Итог)Офицеры.xlsx')
    
    elif number==2:
        name_column=[
        "ФГОО ВО в котором обучается студент",#1
        "ФГОО ВО при которой создан ВУЦ",#2
        "ВУС",#3 
        "Код должности",
        "ОВУ",#4
        "ФГОС",#5
        "Программа военной подготовки",#6
        "Год зачисления в ВУЗ",#7
        "Год начала обучения в ВУЦ",#8
        "Месяц нала обучения в ВУЦ",#9
        "Срок обучения (месяцев)",#10
        "Год окончания ВУЦ",#11
        "Месяц окончания обучения в ВУЦ",#12
        "Отметка о завершении военной подготовки",#13
        "Год окончания ВУЗа",#14
        "Месяц окончания обучения в ВУЗе",#15
        "Личный номер",#16
        "Фамилия",#17
        "Имя",#18
        "Отчество",#19
        "Дата рождения",#20
        "Место рождения",#21
        "Национальность",#22
        "Пол",#23
        "СНИЛС",#24
        "ИНН",#25
        "серия1",#26
        "серия2",#27
        "Номер",#28
        "Выдан",#29
        "Код подразделения",#30
        "Кем выдан",#31
        "Адрес регистрации",#32
        "Номер телефона",#33
        "Семейное положение",#34
        "Количество детей",#35
        "Военный комиссариат (по месту жительства)",#36
        "Военный комиссариат (в который будет направлено личное дело)",#37
        "По призыву",#38
        "По контракту",#39
        "Примечание",#40
        "Статус",	#41
        "Приказ о зачислении",	#42
        "Приказ о отчислении",	#43
        "Причина отчисления"    #44			
        ]
        shutil.copyfile('workprogram/'+'Не трогать!(Шаблон солдат).xlsx', 'Итоговые таблицы/(Итог)Солдаты.xlsx')
        wb = op.load_workbook('Итоговые таблицы/(Итог)Солдаты.xlsx', data_only=True)
        sheet = wb.active
        id=0
        for j in range (len( information)):
            vuc=str(information[j].get('ВУС'))            
            program_pdgotovki=str(information[j].get('Программа военной подготовки'))
            if  program_pdgotovki[:2]=='со' or program_pdgotovki[:2]=='ря' or vuc=='220' or vuc=='233' or vuc=='250' or vuc=='262'  :
                    for i in range (1,len(name_column)+1):
                        tmp=''
                        if  name_column[i-1] in information[j].keys(): tmp=information[j][name_column[i-1]]
                        sheet.cell(row=id+5, column=i, value=tmp)
                    id+=1
        wb.save('Итоговые таблицы/(Итог)Солдаты.xlsx')
    elif number==3:
        print('своя таблица')
    else:
        name_column=[]
        for elem in information:
            name_column=name_column+list(elem.keys())
        name_column=list(set(name_column))
       
        name_column.remove('Фамилия')
        name_column.remove('Имя')
        name_column.remove('Отчество')
        name_column.insert(0, 'Фамилия')
        name_column.insert(1, 'Имя')
        name_column.insert(2, 'Отчество')


        wb = Workbook()
        ws = wb.active

        #Код, который выводит в первый столбец инфу с первого массива. Нужно сделать +- то же самое для остальных и все по сути.
        for i in range (1,len(name_column)+1):
            ws.cell(row=1, column=i, value=name_column[i-1])

        for j in range (len( information)):
            for i in range (1,len(name_column)+1):
                tmp=''
                if  name_column[i-1] in information[j].keys(): tmp=information[j][name_column[i-1]]
                ws.cell(row=j+2, column=i, value=tmp)

        wb.save("Итоговые таблицы/Все данные.xlsx")

if __name__ == "__main__":
  #  try:
        all_files=listdir(getcwd()+'/Таблицы откуда берём информацию')
        excel_name=[]
        for elem in all_files:
            filename, file_extension = path.splitext(elem)
            if  file_extension=='.xlsx': excel_name.append(elem)
        

        print("Программа увидела следующие таблицы:")
        for elem in excel_name:
            print(elem)
        chose=input("------------------------------------------------------------------------------\nПродожить? \n1)Да ")
        if chose=='1':
            information_list=[]
            for elem in excel_name:
                print('\nЧитаем файл   '+str(elem))
                information_list.append(read_excel('Таблицы откуда берём информацию/'+elem))
            print ("\n" * 100)
            print('Данные успешно собраны')
            print('Совмещаем данные')
            
            
            if len(information_list)==1:
                tmp=uniq_list(information_list[0])
                information_list.append(tmp)
                print(len(information_list))
                information_list.pop(0)
            
            while len(information_list)>1:
                tmp=sum(information_list[0],information_list[1])
                information_list.pop(1)
                information_list.pop(0)

                information_list.append(tmp)
            print ("\n" * 100)
            print('Данные совмещены')
            while True:
                tmp=input('Какую таблицу создаём?\n1)Две таблицы: Солдаты, Офицеры\n2)Выгрузить все данные в произвольную таблицу ')
                if tmp=='1':
                    write_excel(information_list[0],1)
                    write_excel(information_list[0],2)
                    input('Успешно, создали "(Итог)Офицеры.xlsx" и "(Итог)Солдаты.xlsx"  в папке "Итоговые таблицы"')
                elif tmp=='2':
                    write_excel(information_list[0],4)
                    input('Успешно,  создали "Все данные.xlsx"  в папке "Итоговые таблицы"')
                else : input('Нет такого варианта..')
  #  except BaseException as exc:
 #      input('Критическая оишбка: Сообщите разработчику исходники и тип оiибки:'+ str(exc))




    # #read_officer читает
    # #СОЕДИНЯЕТ содинение+ проверка на уникальность 
    # #записывает в файл
    # files = listdir(".")
    # excel_name=[]
    # for elem in files:
    #     excel_name.append(elem)
    # print (excel_name)

    # tmp1=read_excel('солдаты.xlsx')

    
    # with open("tmp.txt", mode="w+") as file:
    #     # tmp1=read_cicle('солдаты.xlsx',True)
    #     tmp2=read_cicle('5 цикл офиц Морозов2 (1).xlsx',False)
    #     file.write(str(write_excel(sum(tmp1,tmp2),2)))


    # list1=uniq_list(read_excel( 'офиц.xlsx'))
    # with open("tmp.txt", mode="w+") as file:
    #     file.write(str(list1))

    #print(list1)
    # list1=[{"Фамилия":'Морозов',"Имя":"Сергей","Отчество":"Сергеевич","Дата рождения":"06-03-2003","ср балл":"1222222222222222222",}]
    # list2=[{"Фамилия":'Морозов',"Имя":"Сергей","Отчество":"Сергеевич","Дата рождения":"06-03-2003","Оценка":"5",},{"Фамилия":'Морозов',"Имя":"Сергей","Отчество":"Сергеевич","Дата рождения":"06-03-2003","Оценка":"5","ср балл":"1222222222222222222"}]
    # print(uniq_list(list2))
